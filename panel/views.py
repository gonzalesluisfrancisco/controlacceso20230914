from django.core import serializers
from django.http.request import HttpRequest
from django.http.response import HttpResponse
from django.shortcuts import render, redirect
from django.http import JsonResponse

from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth import get_user_model
from django.utils.encoding import force_str

from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes

from . import models

# from datetime import datetime
from django.utils import timezone
from django.utils.timezone import activate
import pytz

from django.db.models import Q

from . import Funciones as Fun
from .Funciones import debugPrint
from django.core.exceptions import ObjectDoesNotExist

import re
import random
import string
from datetime import datetime
import traceback
import openpyxl

import json

from django.db.models import Avg, F, ExpressionWrapper, fields, Func
from django.db.models import Count

# Create your views here.
TEMPLATE_DIRS = (
    'os.path.join(BASE_DIR, "templates")'
)


def home_view(request: HttpRequest) -> HttpResponse:
    return render(request, 'home.html')


def index_noautenticado(request):

    return render(request, "index.html")


def autenticacion(request):
    if request.method == "POST":

        if request.POST.get("Comando") == "VerificarLogin":
            Usuario = request.POST.get("Usuario")
            Contrasena = request.POST.get("Contrasena")
            Captcha = request.POST.get("Captcha")
            FormatoValido, Mensaje = Fun.FormatoLoginValidos(
                Usuario, Contrasena, Captcha)
            if not FormatoValido:
                return JsonResponse({"Estado": "Invalido", 'Mensaje': Mensaje})
            DatosValidos, Mensaje = Fun.DatosLoginValidos(
                request, Usuario, Contrasena)
            if DatosValidos:
                Correo = Fun.EnviarToken(Usuario, Contrasena, request)
                return JsonResponse({"Estado": "Valido", 'Mensaje': Mensaje, "Correo": Correo})
            return JsonResponse({"Estado": "Invalido", 'Mensaje': Mensaje})

        elif request.POST.get("Comando") == "VerificarToken":
            Usuario = request.POST.get("Usuario")
            Contrasena = request.POST.get("Contrasena")
            Token = request.POST.get("Token")
            if len(Token) != 6:
                return JsonResponse({"Estado": "Invalido", "Mensaje": "El Token debe ser de 8 digitos"})
            if not Fun.VerificarToken(Token):
                return JsonResponse({"Estado": "Invalido", 'Mensaje': "El token ingresado no es correcto"})
            DatosValidos, Mensaje = Fun.DatosLoginValidos(
                request, Usuario, Contrasena)
            if not DatosValidos:
                return JsonResponse({"Estado": "Invalido", 'Mensaje': "No deberias poder ver esto"})
            login(request, authenticate(
                request, username=Usuario, password=Contrasena))
            return JsonResponse({"Estado": "Valido", "Mensaje": "El token ha sido validado correctamente"})

        elif request.POST.get("Comando") == "RecuperarCuenta":
            Correo = request.POST.get("Correo")
            if len(Correo) == 0:
                return JsonResponse({"Estado": "Invalido", "Mensaje": "Debe ingresar su correo electronico"})
            try:
                Usuario = User.objects.get(email=Correo)
            except ObjectDoesNotExist:
                return JsonResponse({"Estado": "Invalido", "Mensaje": "El correo ingresado no se encuentra registrado"})
            token = default_token_generator.make_token(Usuario)
            uid = urlsafe_base64_encode(force_bytes(Usuario.pk))
            current_site = get_current_site(request)
            Asunto = 'Recuperar contraseña - DIACSA'
            reset_url = f"https://{current_site}/reset-password/{uid}/{token}/"
            MensajeHTML = f"""\
            <html>
            <head></head>
            <body>
                <p>Hola, <span style="font-size: larger;">{Usuario.first_name}</span>!</p>
                <p>Entra a este enlace para recuperar tu contraseña: <br><span style="font-size: larger;"><b>{reset_url}</b></span></p>
            </body>
            </html>
            """
            Fun.EnviaCorreo(Usuario.email, Asunto, MensajeHTML)
            return JsonResponse({"Estado": "Valido", "Mensaje": "El enlace de recuperacion ha sido enviado a su correo"})

        else:
            return JsonResponse({"Estado": "Invalido", 'Mensaje': "El comando es desconocido"})

    signout(request)
    return render(request, "login.html")


def reset_password(request, uidb64, token):
    if request.method == 'POST':
        try:
            UID = force_str(urlsafe_base64_decode(uidb64))
            Usuario = User.objects.get(pk=UID)
        except (TypeError, ValueError, OverflowError, User.DoesNotExist):
            Usuario = None
        if Usuario is None or not default_token_generator.check_token(Usuario, token):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El enlace de recuperacion no es valido"})
        NuevaContrasena = request.POST.get('Contrasena1')
        ConfirmacionContrasena = request.POST.get('Contrasena2')
        if NuevaContrasena != ConfirmacionContrasena:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "Las contraseñas ingresadas no coinciden"})
        if not Fun.ContrasenaEsFuerte(NuevaContrasena):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "La nueva contraseña debe tener:\nUna mayuscula.\nUna minuscula.\nSin numeros consecutivos.\nMinimo 8 caracteres"})
        Usuario.set_password(NuevaContrasena)
        Usuario.save()
        return JsonResponse({"Estado": "Valido", "Mensaje": "La contraseña se ha cambiado correctamente"})
    else:
        return render(request, 'reset_password.html')


@login_required(login_url='autenticacion')
def signout(request):
    logout(request)
    return redirect("autenticacion")


@login_required(login_url='autenticacion')
def index(request):
    debugPrint("Ingreso a intentar leer el livedata")
    users = models.LiveData.objects.all()
    data = {'total': users.count()}
    return render(request, "index.html", data)

def sacaHoraPromedio(registros):
    for registro in registros:
        debugPrint(registro.id)
        debugPrint(registro.f_evento)        
        debugPrint(registro.h_evento)
    return ["1:1:1","2:2:2","3:3:3","4:4:4","5:5:5","6:6:6","7:7:7"]

@login_required(login_url='autenticacion')
def dashboard(request):
    if request.method == "POST":
        debugPrint("POST en filtrar datos")
        debugPrint("Respuesta al POST OK")
        areaFiltro = request.POST.get("area", "")
        servicioFiltro = request.POST.get("servicio", "")
        cargoFiltro = request.POST.get("cargo", "")
        fechaFiltro = request.POST.get("fecha", "")
        año = fechaFiltro[0:4]
        mes = fechaFiltro[5:7]

        debugPrint(areaFiltro)
        debugPrint(servicioFiltro)
        debugPrint(cargoFiltro)
        debugPrint(fechaFiltro)

        registros = models.Historial.objects.filter(~(Q(status="10")|Q(status="11")))

        if(areaFiltro != ""):
            registros = registros.filter(area=areaFiltro)
        if(servicioFiltro != ""):
            registros = registros.filter(servicio=servicioFiltro)
        if(cargoFiltro != ""):
            registros = registros.filter(cargo=cargoFiltro)
        if(fechaFiltro != ""):
            registros = registros.filter(Q(f_evento__icontains=fechaFiltro))
        
        cantidadIngreso = registros.filter(evento='Ingreso').count()
        cantidadSalida  = registros.filter(evento='Salida').count()
        debugPrint(cantidadIngreso)
        #Arma datos tabla 1
        #[["PERSONAL QUE INGRESO", cantidadIngreso, 'red'],["PERSONAL QUE SALIO", cantidadSalida, 'blue']]
        values = [["PERSONAL QUE INGRESO", cantidadIngreso, 'red']]
        values.append(["PERSONAL QUE SALIO", cantidadSalida, 'blue'])
        values_JSON1 = json.dumps(values)
        #Arma datos tabla 2
            #Hacer un get con la todos los objetos del mes
            ##[
            ## ["Lunes", cantidadIngreso, cantidadSalidas],
            ## ["Martes", cantidadIngreso, cantidadSalidas],
            ## ["Miercoles", cantidadIngreso, cantidadSalidas],
            ## ["Jueves", cantidadIngreso, cantidadSalidas],
            ## ["Viernes", cantidadIngreso, cantidadSalidas],
            ## ["Sabado", cantidadIngreso, cantidadSalidas],
            ## ["Domingo", cantidadIngreso, cantidadSalidas],
            ## ]
        values = [["Lunes", 1, 1],
             ["Martes", 2, 2],
             ["Miercoles", 3, 3],
             ["Jueves", 4, 4],
             ["Viernes", 5, 5],
             ["Sabado", 6, 6],
             ["Domingo", 7, 7],
             ]
        
        values_JSON2 = json.dumps(values)
        #Arma datos tabla 3
        ##[
            ## ["Lunes", HoraPromedioIngreso, HoraPromedioSalidas],
            ## ["Martes", HoraPromedioIngreso, HoraPromedioSalidas],
            ## ["Miercoles", HoraPromedioIngreso, HoraPromedioSalidas],
            ## ["Jueves", HoraPromedioIngreso, HoraPromedioSalidas],
            ## ["Viernes", HoraPromedioIngreso, HoraPromedioSalidas],
            ## ["Sabado", HoraPromedioIngreso, HoraPromedioSalidas],
            ## ["Domingo", HoraPromedioIngreso, HoraPromedioSalidas],
            ## ]
        values = [
             ["Lunes", [1,1,1], [1,1,1]],
             ["Martes", [2,2,2], [2,2,2]],
             ["Miercoles", [3,3,3], [3,3,3]],
             ["Jueves", [4,4,4], [4,4,4]],
             ["Viernes", [5,5,5], [5,5,5]],
             ["Sabado", [6,6,6], [6,6,6]],
             ["Domingo", [7,7,7], [7,7,7]],
             ]
        eventos_Ingreso = registros.filter(evento='Ingreso')
        horaIngreso = sacaHoraPromedio(eventos_Ingreso)
        eventos_Salida = registros.filter(evento='Salida')
        horaSalida = sacaHoraPromedio(eventos_Salida)
        horas_semana = [[horaIngreso[0],horaSalida[0]],
                        [horaIngreso[1],horaSalida[1]],
                        [horaIngreso[2],horaSalida[2]],
                        [horaIngreso[3],horaSalida[3]],
                        [horaIngreso[4],horaSalida[4]],
                        [horaIngreso[5],horaSalida[5]],
                        [horaIngreso[6],horaSalida[6]],
                        ]
        print(horas_semana)
        values_JSON3 = json.dumps(values)

        return JsonResponse({"Estado": "OK", "datos": "Datos para las tablas", 'valuesTabla1': values_JSON1, 'valuesTabla2': values_JSON2, 'valuesTabla3': values_JSON3})
    elif request.method == "GET":
        debugPrint("Ingreso a dashboard")
        areas = models.Historial.objects.values(
            'area').distinct().order_by("-area")
        servicios = models.Historial.objects.values(
            'servicio').distinct().order_by("-servicio")
        cargos = models.Historial.objects.values(
            'cargo').distinct().order_by("-cargo")
        guardias = models.Historial.objects.values(
            'guardia').distinct().order_by("-guardia")
        # X_var = 'Personal'
        # Y_var = 'Ingreso/Salidas'
        # # Color_var = {role: 'style'}
        # # values = [[X_var, Y_var]]
        # # debugPrint(models.LiveData.objects.count())
        # values = [["PERSONAL QUE INGRESO",
        #            models.LiveData.objects.count(), 'red']]
        # values.append(["PERSONAL QUE SALIO", 5, 'blue'])
        # values_JSON = json.dumps(values)
        values_JSON = ""
        return render(request, "dashboard/dashboard.html", {'values': values_JSON, 'filtrosArea': areas, 'filtrosServicio': servicios, 'filtroCargo': cargos, 'filtroGuardia': guardias})


@login_required(login_url='autenticacion')
def listar(request):
    if request.method == "POST" and request.user.is_superuser and (request.POST.get("Comando") == "Agregar"):
        debugPrint("Entra a Agregar Plantilla por excel")
        header = ["N° PERSONA", "APELLIDO PATERNO", "APELLIDO MATERNO", "NOMBRE", "DNI", "FECHA DE NACIMIENTO", "PROYECTO", "CENTRO DE COSTE", "TIPO DE TRABAJADOR",
                  "CLAVE DE SEXO", "FECHA DE ALTA", "FECHA DE BAJA", "MOTIVO DE CESE", "CARGO", "CÓDIGO DE TARJETA", "ÁREA", "SERVICIO", "SUPERVISIÓN", "GUARDIA", "CORREO", "N° CELULAR"]
        n_columnas = len(header)
        ExcelFile = request.FILES.get("excelFile")
        if not ExcelFile:
            return JsonResponse({'Estado': 'Invalido', 'Mensaje': "No se ha podido leer el archivo"})
        try:
            Excel = openpyxl.load_workbook(
                ExcelFile, read_only=True, keep_vba=True, data_only=True, keep_links=False)
        except Exception as e:
            return JsonResponse({'Estado': 'Invalido', 'Mensaje': str(e) + "\n" + traceback.format_exc()})
        Hoja = Excel.active
        if not Hoja:
            return JsonResponse({'Estado': 'Invalido', 'Mensaje': "El archivo no tienen ninguna hoja"})
        FilasTotales = Hoja.max_row
        debugPrint(f"Filas del excel: {FilasTotales}")
        debugPrint("La integridad del archivo es correcta")
        for i in range(0, n_columnas):
            try:
                if not str(Hoja[1][i].value).strip().upper() == header[i].strip():
                    debugPrint(
                        f"{Hoja[1][i].value} != {header[i]}")
                    return JsonResponse({'Mensaje': 'La cabecera es incorrecta (parcialmente)', 'Estado': "Invalido"})
            except:
                return JsonResponse({'Mensaje': 'La cabecera es incorrecta', 'Estado': "Invalido"})
            debugPrint(
                f"Cabecera {i}, correcta: {Hoja[1][i].value}")
        debugPrint("La cabecera es correcta")
        filasInvalidas = []
        motivoInvalido = []
        # if FilasTotales == None:
        #     return JsonResponse({'Mensaje': "No hay filas para ser procesadas", 'Estado': "Invalido"})
        for i in range(2, FilasTotales):
            Fila = Hoja[i]
            for j in range(0, n_columnas):
                debugPrint(f"{header[j]}: {Fila[j].value}")
            correo = str(Fila[19].value).strip()
            if correo == "None":
                correo = ""
            motivo_cese = str(Fila[12].value).strip()
            if motivo_cese == "None":
                motivo_cese = ""
            telefono = str(Fila[20].value).strip()
            if telefono == "None":
                telefono = ""

            if len(str(Fila[0].value)) != 8 or Fila[0].value == None or Fila[0].value.strip() == "":
                filasInvalidas.append(i)
                motivoInvalido.append("Nro de persona no tiene  8 digitos")
                continue
            if Fila[1].value == None or Fila[1].value.strip() == "":
                filasInvalidas.append(i)
                motivoInvalido.append("Apellido paterno vacio")
                continue
            if Fila[2].value == None or Fila[2].value.strip() == "":
                filasInvalidas.append(i)
                motivoInvalido.append("Apellido materno vacio")
                continue
            if Fila[3].value == None or Fila[3].value.strip() == "":
                filasInvalidas.append(i)
                motivoInvalido.append("Nombre vacio")
                continue
            if len(str(Fila[4].value)) != 8 or Fila[4].value == None or str(Fila[4].value).strip() == "":
                filasInvalidas.append(i)
                motivoInvalido.append("DNI no tiene 8 digitos")
                continue
            if not isinstance(Fila[5].value, datetime):
                filasInvalidas.append(i)
                motivoInvalido.append("Fecha de nacimiento invalida")
                continue
            if Fila[6].value.strip().upper() != "UND. SAN RAFAEL" and Fila[6].value.strip().upper() != "UND. SAN CRISTÓBAL" and Fila[6].value.strip().upper() != "UND. RAURA" and Fila[6].value.strip().upper() != "UND. CHUNGAR":
                filasInvalidas.append(i)
                motivoInvalido.append("Proyecto invalido")
                continue
            if len(str(Fila[7].value)) != 10:
                filasInvalidas.append(i)
                motivoInvalido.append("Centro de coste no tiene 10 digitos")
                continue
            if Fila[8].value.strip().upper() != "OBRERO" and Fila[8].value.strip().upper() != "PRACTICANTE" and Fila[8].value.strip().upper() != "GERENTE" and Fila[8].value.strip().upper() != "EMPLEADO":
                filasInvalidas.append(i)
                motivoInvalido.append("Tipo de trabajador invalido")
                continue
            if Fila[9].value.strip().upper() != "M" and Fila[9].value.strip().upper() != "F":
                filasInvalidas.append(i)
                motivoInvalido.append("Clave de sexo invalido")
                continue
            if not isinstance(Fila[10].value, datetime):
                filasInvalidas.append(i)
                motivoInvalido.append("Fecha de alta invalido")
                continue
            if not isinstance(Fila[11].value, datetime) and (Fila[11].value != None and Fila[11].value.strip().upper() != ""):
                filasInvalidas.append(i)
                motivoInvalido.append("Fecha de baja invalido")
                continue
            # Columna 12 no requiere validacion
            if Fila[13].value == None or Fila[13].value.strip().upper() == "":
                filasInvalidas.append(i)
                motivoInvalido.append("Cargo vacio")
                continue
            if len(str(Fila[14].value)) != 8 or Fila[14].value == None or Fila[14].value.strip() == "":
                filasInvalidas.append(i)
                motivoInvalido.append("ID de tarjeta invalido")
                continue
            if Fila[15].value.strip().upper() != "GESTIÓN HUMANA" and Fila[15].value.strip().upper() != "MANTENIMIENTO" and Fila[15].value.strip().upper() != "OPERACIONES" and Fila[15].value.strip().upper() != "SERVICIOS AUXILIARES MINA" and Fila[15].value.strip().upper() != "LOGÍSTICA" and Fila[15].value.strip().upper() != "ORE SORTING" and Fila[15].value.strip().upper() != "SHOTCRETE" and Fila[15].value.strip().upper() != "SSMA" and Fila[15].value.strip().upper() != "OFICINA TÉCNICA":
                filasInvalidas.append(i)
                motivoInvalido.append("Area invalida")
                continue
            if str(Fila[16].value).strip().upper() == "" or Fila[16].value == None:
                filasInvalidas.append(i)
                motivoInvalido.append("Servicio invalida")
                continue
            # Columna 16 tiene muchos items por validar
            if str(Fila[17].value).strip().upper() != "LINEA SUPERVISORA" and str(Fila[17].value).strip().upper() != "LINEA OBRERA":
                filasInvalidas.append(i)
                motivoInvalido.append("Supervision invalido")
                continue
            if Fila[18].value.strip().upper() != "A" and Fila[18].value.strip().upper() != "B" and Fila[18].value.strip().upper() != "C":
                filasInvalidas.append(i)
                motivoInvalido.append("Guardia invalida")
                continue
            if Fila[19].value != None and str(Fila[19].value).strip().upper() != "" and not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', str(Fila[19].value).strip()):
                filasInvalidas.append(i)
                motivoInvalido.append("Correo invalido")
                continue
            if len(str(Fila[20].value).strip()) != 9 and (Fila[20].value != None and str(Fila[20].value).strip() != ""):
                filasInvalidas.append(i)
                motivoInvalido.append("Numero de celular no tiene 9 digitos")
                continue
            debugPrint(f"Fila {i} correcta.")
            nuevoPersonal = models.PersonalRegistrado()
            nuevoPersonal.n_persona = str(Fila[0].value).strip()
            nuevoPersonal.ap_paterno = str(Fila[1].value).strip()
            nuevoPersonal.ap_materno = str(Fila[2].value).strip()
            nuevoPersonal.nombre = str(Fila[3].value).strip()
            nuevoPersonal.dni = str(Fila[4].value).strip()
            nuevoPersonal.f_nac = datetime.strptime(
                str(Fila[5].value).strip(), "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
            nuevoPersonal.proyecto = str(Fila[6].value).strip()
            nuevoPersonal.centro_coste = str(Fila[7].value).strip()
            nuevoPersonal.tipo_trabajador = str(Fila[8].value).strip()
            nuevoPersonal.clave_sexo = str(Fila[9].value).strip()
            nuevoPersonal.f_alta = datetime.strptime(
                str(Fila[10].value).strip(), "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
            if re.match(r"^\d{4}-\d{2}-\d{2}$", str(Fila[11].value)):
                nuevoPersonal.f_baja = datetime.strptime(
                    str(Fila[11].value).strip(), "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
            nuevoPersonal.motivo_cese = motivo_cese
            nuevoPersonal.cargo = str(Fila[13].value).strip()
            nuevoPersonal.card_id = str(Fila[14].value).strip()
            nuevoPersonal.area = str(Fila[15].value).strip()
            nuevoPersonal.servicio = str(Fila[16].value).strip()
            nuevoPersonal.supervision = str(Fila[17].value).strip()
            nuevoPersonal.guardia = str(Fila[18].value).strip()
            nuevoPersonal.correo = correo
            nuevoPersonal.n_celular = telefono
            nuevoPersonal.save()
        debugPrint(len(filasInvalidas))
        if len(filasInvalidas) != 0:
            txt = ""
            idx = 0
            for i in filasInvalidas:
                debugPrint(f"Fila {i}: {motivoInvalido[idx]}")
                txt = f"{txt}Fila {i}: {motivoInvalido[idx]}<br>"
                idx += 1
            return JsonResponse({'Mensaje': f'Planilla actualizada correctamente<br>Las siguientes filas no han podido ser procesadas:<br>{txt}', 'Estado': "Valido"})
        return JsonResponse({'Mensaje': 'Planilla actualizada correctamente', 'Estado': "Valido"})
    if request.method == "POST" and request.user.is_superuser and (request.POST.get("Comando") == "Agregar"):
        ExcelFile = request.FILES.get("excelFile")
        if ExcelFile:
            try:
                Excel = openpyxl.load_workbook(
                    ExcelFile, read_only=True, keep_vba=True, data_only=True, keep_links=False)
                Hoja = Excel.active
                FilasTotales = Hoja.max_row
                debugPrint("Llego a leer el excel, eliminar")
                for i in range(0, n_columnas):
                    try:
                        if not str(Hoja[1][i].value).strip().upper() == header[i].strip():
                            debugPrint(
                                f"{Hoja[1][i].value} != {header[i]}")
                            return JsonResponse({'Mensaje': 'La cabecera es incorrecta (parcialmente)', 'Estado': "Invalido"})
                    except:
                        return JsonResponse({'Mensaje': 'La cabecera es incorrecta', 'Estado': "Invalido"})
                    debugPrint(
                        f"Cabecera {i}, correcta: {Hoja[1][i].value}")
                debugPrint("Hasta aqui cabecera correcta")
                for i in range(2, FilasTotales):
                    None
            except Exception as e:
                return JsonResponse({'Estado': 'Invalido', 'Mensaje': str(e) + "\n" + traceback.format_exc()})
        debugPrint("El archivo no se ha podido leer")
        return JsonResponse({"Estado": "Invalido", "Mensaje": "El archivo no se ha podido leer"})

    if request.method == "POST" and request.POST.get("Comando") == "eliminarPlantilla":
        if not request.user.is_superuser and not request.user.is_staff:
            return JsonResponse({"Mensaje": "Su cuenta no tiene los privilegios para realizar esta accion"})
        tablaPlantilla = models.PersonalRegistrado.objects.all()
        tablaPlantilla.delete()
        return JsonResponse({"Mensaje": "La plantilla ha sido borrada correctamente", "Estado": "Valido"})
    users = models.PersonalRegistrado.objects.all()
    datos = {'personalregistrado': users}
    if request.user.is_staff:
        return render(request, "crud_aesadiacsa/listar.html", datos)
    return render(request, "crud_aesadiacsa/listar_basic.html", datos)


@login_required(login_url='autenticacion')
def agregar(request):
    if request.method == 'POST' and request.POST.get('comando') == 'agregarManual':
        debugPrint("Entra a agregar personal")
        n_persona = request.POST.get('n_persona', '')
        nombre = request.POST.get('nombre', '')
        ap_paterno = request.POST.get('ap_paterno', '')
        ap_materno = request.POST.get('ap_materno', '')
        dni = request.POST.get('dni', '')
        f_nac = request.POST.get('f_nac', '')
        proyecto = request.POST.get('proyecto', '')
        centro_coste = request.POST.get('centro_coste', '')
        tipo_trabajador = request.POST.get('tipo_trabajador', '')
        clave_sexo = request.POST.get('clave_sexo', '')
        f_alta = request.POST.get('f_alta', '')
        f_baja = request.POST.get('f_baja', '')
        motivo_cese = request.POST.get('motivo_cese', '')
        cargo = request.POST.get('cargo', '')
        card_id = request.POST.get('card_id', '').upper()
        area = request.POST.get('area', '').upper()
        servicio = request.POST.get('servicio', '')
        supervision = request.POST.get('supervision', '').upper()
        guardia = request.POST.get('guardia', '')
        correo = request.POST.get('correo', '')
        telefono = request.POST.get('telefono', '')
        debugPrint(request.POST.get('n_persona', ''))
        debugPrint(request.POST.get('ap_paterno', ''))
        debugPrint(request.POST.get('ap_materno', ''))
        debugPrint(request.POST.get('nombre', ''))
        debugPrint(request.POST.get('dni', ''))
        debugPrint(request.POST.get('f_nac', ''))
        debugPrint(request.POST.get('proyecto', ''))
        debugPrint(request.POST.get('centro_coste', ''))
        debugPrint(request.POST.get('tipo_trabajador', ''))
        debugPrint(request.POST.get('clave_sexo', ''))
        debugPrint(request.POST.get('f_alta', ''))
        debugPrint(request.POST.get('f_baja', ''))
        debugPrint(request.POST.get('motivo_cese', ''))
        debugPrint(request.POST.get('cargo', ''))
        debugPrint(request.POST.get('card_id', ''))
        debugPrint(request.POST.get('area', '').upper())
        debugPrint(request.POST.get('servicio', ''))
        debugPrint(request.POST.get('supervision', ''))
        debugPrint(request.POST.get('guardia', ''))
        debugPrint(request.POST.get('correo', ''))
        debugPrint(request.POST.get('n_celular', ''))
        debugPrint("Entra a chekear datos")
        if len(n_persona) != 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El numero de persona debe tener 8 digitos"})
        if len(ap_paterno) == 0:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El apellido paterno no puede estar en blanco"})
        if len(ap_materno) == 0:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El apellido materno no puede estar en blanco"})
        if len(nombre) == 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El nombre no puede estar en blanco"})
        if len(dni) != 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El numero de dni debe tener 8 digitos"})
        if len(f_nac) == 0:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "La fecha de nacimiento no puede esta en blanco"})
        if proyecto.upper() == "UND. CHUNGAR" and proyecto.upper() == "UND. SAN RAFAEL" and proyecto.upper() == "UND. SAN CRISTÓBAL" and proyecto.upper() == "UND. RAURA":
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El proyecto debe ser UND. SAN RAFAEL, UND. SAN CRISTÓBAL o UND. RAURA"})
        if len(centro_coste) != 10:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El centro de coste debe tener 10 digitos"})
        if tipo_trabajador.upper() != "OBRERO" and tipo_trabajador.upper() != "PRACTICANTE" and tipo_trabajador.upper() != "GERENTE" and tipo_trabajador.upper() != "EMPLEADO":
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El tipo de trabajador debe ser OBRERO, PRACTICANTE, GERENTE o EMPLEADO"})
        if clave_sexo.upper() != "M" and clave_sexo.upper() != "F":
            return JsonResponse({"Estado": "Invalido", "Mensaje": "La clave de sexo debe ser M o F"})
        if len(f_alta) == 0:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "La fecha de alta no puede estar en blanco"})
        if len(f_baja) == 0:
            f_baja = None
        if len(cargo) == 0:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El cargo no puede estar vacio"})
        if len(card_id) != 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El card ID debe tener 8 digitos"})
        if area != "GESTIÓN HUMANA" and area != "MANTENIMIENTO" and area != "OPERACIONES" and area != "SERVICIOS AUXILIARES MINA" and area != "LOGÍSTICA" and area != "ORE SORTING" and area != "SHOTCRETE" and area != "SSMA" and area != "OFICINA TÉCNICA":
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El area debe ser GESTIÓN HUMANA, MANTENIMIENTO, OPERACIONES, SERVICIOS AUXILIARES MINA, LOGÍSTICA, ORE SORTING, SHOTCRETE, SSMA u OFICINA TÉCNICA"})
        if len(servicio) == 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El servicio no puede estar vacio"})
        if supervision != "LINEA SUPERVISORA" and supervision != "LINEA OBRERA":
            return JsonResponse({"Estado": "Invalido", "Mensaje": "La supervision debe ser LINEA SUPERVISORA o LINEA OBRERA"})
        if len(guardia) != 1:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "La guardia debe ser A, B o C"})
        usuario = models.PersonalRegistrado()
        usuario.id = models.PersonalRegistrado.objects.count() + 1
        usuario.n_persona = n_persona
        usuario.ap_paterno = ap_paterno
        usuario.ap_materno = ap_materno
        usuario.nombre = nombre
        usuario.dni = dni
        usuario.f_nac = f_nac
        usuario.proyecto = proyecto
        usuario.centro_coste = centro_coste
        usuario.tipo_trabajador = tipo_trabajador
        usuario.clave_sexo = clave_sexo
        usuario.f_alta = f_alta
        usuario.f_baja = f_baja
        usuario.motivo_cese = motivo_cese
        usuario.cargo = cargo
        usuario.card_id = card_id
        usuario.area = area
        usuario.servicio = servicio
        usuario.supervision = supervision
        usuario.guardia = guardia
        usuario.correo = correo
        usuario.n_celular = telefono
        usuario.save()
        return JsonResponse({"Mensaje": "Los cambios se guardaron correctamente"})
    else:
        return render(request, "crud_aesadiacsa/agregar.html")


@login_required(login_url='autenticacion')
def actualizar(request, codigo):
    if request.method == 'POST' and request.POST.get("comando") == "actualizarPersonal":
        debugPrint("Entra a actualizar personal")
        id = request.POST.get('id', '')
        n_persona = request.POST.get('n_persona', '')
        nombre = request.POST.get('nombre', '')
        ap_paterno = request.POST.get('ap_paterno', '')
        ap_materno = request.POST.get('ap_materno', '')
        dni = request.POST.get('dni', '')
        f_nac = request.POST.get('f_nac', '')
        proyecto = request.POST.get('proyecto', '')
        centro_coste = request.POST.get('centro_coste', '')
        tipo_trabajador = request.POST.get('tipo_trabajador', '')
        clave_sexo = request.POST.get('clave_sexo', '')
        f_alta = request.POST.get('f_alta', '')
        f_baja = request.POST.get('f_baja', '')
        motivo_cese = request.POST.get('motivo_cese', '')
        cargo = request.POST.get('cargo', '')
        card_id = request.POST.get('card_id', '').upper()
        area = request.POST.get('area', '').upper()
        servicio = request.POST.get('servicio', '')
        supervision = request.POST.get('supervision', '').upper()
        guardia = request.POST.get('guardia', '')
        correo = request.POST.get('correo', '')
        telefono = request.POST.get('telefono', '')
        debugPrint(request.POST.get('n_persona', ''))
        debugPrint(request.POST.get('ap_paterno', ''))
        debugPrint(request.POST.get('ap_materno', ''))
        debugPrint(request.POST.get('nombre', ''))
        debugPrint(request.POST.get('dni', ''))
        debugPrint(request.POST.get('f_nac', ''))
        debugPrint(request.POST.get('proyecto', ''))
        debugPrint(request.POST.get('centro_coste', ''))
        debugPrint(request.POST.get('tipo_trabajador', ''))
        debugPrint(request.POST.get('clave_sexo', ''))
        debugPrint(request.POST.get('f_alta', ''))
        debugPrint(request.POST.get('f_baja', ''))
        debugPrint(request.POST.get('motivo_cese', ''))
        debugPrint(request.POST.get('cargo', ''))
        debugPrint(request.POST.get('card_id', ''))
        debugPrint(request.POST.get('area', '').upper())
        debugPrint(request.POST.get('servicio', ''))
        debugPrint(request.POST.get('supervision', ''))
        debugPrint(request.POST.get('guardia', ''))
        debugPrint(request.POST.get('correo', ''))
        debugPrint(request.POST.get('n_celular', ''))
        debugPrint("Entra a chekear datos")
        if len(id) == 0:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "Algo salio mal, actualize la pagina e intente nuevamente"})
        if len(n_persona) != 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El numero de persona debe tener 8 digitos"})
        if len(ap_paterno) == 0:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El apellido paterno no puede estar en blanco"})
        if len(ap_materno) == 0:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El apellido materno no puede estar en blanco"})
        if len(nombre) == 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El nombre no puede estar en blanco"})
        if len(dni) != 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El numero de dni debe tener 8 digitos"})
        if len(f_nac) == 0:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "La fecha de nacimiento no puede esta en blanco"})
        if proyecto.upper() == "UND. CHUNGAR" and proyecto.upper() == "UND. SAN RAFAEL" and proyecto.upper() == "UND. SAN CRISTÓBAL" and proyecto.upper() == "UND. RAURA":
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El proyecto debe ser UND. SAN RAFAEL, UND. SAN CRISTÓBAL o UND. RAURA"})
        if len(centro_coste) != 10:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El centro de coste debe tener 10 digitos"})
        if tipo_trabajador.upper() != "OBRERO" and tipo_trabajador.upper() != "PRACTICANTE" and tipo_trabajador.upper() != "GERENTE" and tipo_trabajador.upper() != "EMPLEADO":
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El tipo de trabajador debe ser OBRERO, PRACTICANTE, GERENTE o EMPLEADO"})
        if clave_sexo.upper() != "M" and clave_sexo.upper() != "F":
            return JsonResponse({"Estado": "Invalido", "Mensaje": "La clave de sexo debe ser M o F"})
        if len(f_alta) == 0:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "La fecha de alta no puede estar en blanco"})
        if len(f_baja) == 0:
            f_baja = None
        if len(cargo) == 0:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El cargo no puede estar vacio"})
        if len(card_id) != 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El card ID debe tener 8 digitos"})
        if area != "GESTIÓN HUMANA" and area != "MANTENIMIENTO" and area != "OPERACIONES" and area != "SERVICIOS AUXILIARES MINA" and area != "LOGÍSTICA" and area != "ORE SORTING" and area != "SHOTCRETE" and area != "SSMA" and area != "OFICINA TÉCNICA":
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El area debe ser GESTIÓN HUMANA, MANTENIMIENTO, OPERACIONES, SERVICIOS AUXILIARES MINA, LOGÍSTICA, ORE SORTING, SHOTCRETE, SSMA u OFICINA TÉCNICA"})
        if len(servicio) == 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El servicio no puede estar vacio"})
        if supervision != "LINEA SUPERVISORA" and supervision != "LINEA OBRERA":
            return JsonResponse({"Estado": "Invalido", "Mensaje": "La supervision debe ser LINEA SUPERVISORA o LINEA OBRERA"})
        if len(guardia) != 1:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "La guardia debe ser A, B o C"})
        usuario = models.PersonalRegistrado.objects.get(id=id)
        usuario.n_persona = n_persona
        usuario.ap_paterno = ap_paterno
        usuario.ap_materno = ap_materno
        usuario.nombre = nombre
        usuario.dni = dni
        usuario.f_nac = f_nac
        usuario.proyecto = proyecto
        usuario.centro_coste = centro_coste
        usuario.tipo_trabajador = tipo_trabajador
        usuario.clave_sexo = clave_sexo
        usuario.f_alta = f_alta
        usuario.f_baja = f_baja
        usuario.motivo_cese = motivo_cese
        usuario.cargo = cargo
        usuario.card_id = card_id
        usuario.area = area
        usuario.servicio = servicio
        usuario.supervision = supervision
        usuario.guardia = guardia
        usuario.correo = correo
        usuario.n_celular = telefono
        usuario.save()
        return JsonResponse({"Mensaje": "Los cambios se guardaron correctamente"})
    else:
        datosuser = models.PersonalRegistrado.objects.get(card_id=codigo)
        debugPrint("Obtuvo datos de usuario")
        debugPrint(datosuser)
        datos = {'personalregistrado': datosuser}

        return render(request, "crud_aesadiacsa/actualizar.html", datos)


@login_required(login_url='autenticacion')
def eliminar(request, codigo):
    tupla = models.PersonalRegistrado.objects.get(card_id=codigo)
    tupla.delete()
    return redirect('listar')


@login_required(login_url='autenticacion')
def livedata(request):
    if request.method == "POST" and request.POST.get("Comando") == "TablaLiveData":
        debugPrint("Tabla LiveData")
        debugPrint("Search")
        draw = int(request.POST.get('draw', 0))
        start = int(request.POST.get('start', 0))
        length = int(request.POST.get('length', 0))
        search_value = request.POST.get('search[value]', '')
        queryset = models.LiveData.objects.order_by(
            "-f_ingreso", "-h_ingreso", "-id")
        if search_value != '':
            queryset = queryset.filter(
                Q(ap_paterno__icontains=search_value) |
                Q(ap_materno__icontains=search_value) |
                Q(nombre__icontains=search_value) |
                Q(dni__icontains=search_value) |
                Q(proyecto__icontains=search_value) |
                Q(cargo__icontains=search_value) |
                Q(area__icontains=search_value) |
                Q(servicio__icontains=search_value) |
                Q(supervision__icontains=search_value) |
                Q(guardia__icontains=search_value)
            ).order_by("-f_ingreso", "-h_ingreso")
        total_records = queryset.count()
        queryset = queryset[start:start+length]
        data = []
        for i, obj in enumerate(queryset, start=0):
            item = {
                'id': str(total_records - start - i),
                'ubicacion': obj.ubicacion,
                'card_id': obj.card_id,
                'nombre': obj.nombre,
                'ap_paterno': obj.ap_paterno,
                'ap_materno': obj.ap_materno,
                'dni': obj.dni,
                'proyecto': obj.proyecto,
                'cargo': obj.cargo,
                'area': obj.area,
                'servicio': obj.servicio,
                'supervision': obj.supervision,
                'guardia': obj.guardia,
                'f_ingreso': obj.f_ingreso,
                'h_ingreso': obj.h_ingreso,
            }
            data.append(item)
        response = {
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": total_records,
            "data": data,
        }
        # debugPrint(response)
        return JsonResponse(response)
    elif request.method == "POST" and request.POST.get("Comando") == "ObtenerHoraTotal":
        debugPrint("ObtenerHoraTotal")
        lima_timezone = pytz.timezone('America/Lima')
        lima_time = timezone.now().astimezone(
            lima_timezone).strftime('%Y-%m-%d %H:%M:%S')
        total = models.LiveData.objects.order_by(
            "-f_ingreso", "-h_ingreso").count()
        return JsonResponse({"Estado": "Valido", "Total": total, "Hora": lima_time})
    # users = models.LiveData.objects.all()
    # activate(pytz.timezone('America/Lima'))
    # debugPrint(timezone.now())
    # datos = { 'livedata' : users,             'fecha_y_hora': timezone.now(),             'total': users.count()}
    return render(request, "livedata/livedata.html")


@login_required(login_url='autenticacion')
def livedata_llenar(request):
    users = models.LiveData.objects.all()
    # users.
    # datosuser = models.PersonalRegistrado.objects.get(id=codigo)
    return redirect('livedata')


@login_required(login_url='autenticacion')
def livedata_agregar(request):
    debugPrint(request.POST.get('comando'))
    if request.method == 'POST' and request.POST.get('comando') == 'consultaDatos':
        debugPrint(request.POST.get('cardid'))
        debugPrint(request.POST.get('nombre'))
        debugPrint(request.POST.get('apellido'))
        debugPrint(request.POST.get('cargo'))
        debugPrint(request.POST.get('f_ingreso'))
        debugPrint(request.POST.get('h_ingreso'))
        # agregar datos
        if len(request.POST.get('cardid', ' ')) != 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El Card ID debe ser de 8 caracteres"})
        elif not request.POST.get('nombre') and not request.POST.get('apellido') and not request.POST.get('cargo') and not request.POST.get('f_ingreso') and not request.POST.get('h_ingreso'):
            try:
                users = models.PersonalRegistrado.objects.get(
                    card_id=request.POST.get('cardid'))
            except ObjectDoesNotExist:
                return JsonResponse({"ubicacion": '',
                                    "nombre": '',
                                     "apellido": '',
                                     "cargo": '',
                                     "f_ingreso": '',
                                     "h_ingreso": ''})
            else:
                return JsonResponse({"ubicacion": '',
                                    "nombre": users.nombre,
                                     "apellido": users.apellido,
                                     "cargo": users.cargo,
                                     "f_ingreso": '',
                                     "h_ingreso": ''})

    if request.method == 'POST' and request.POST.get('comando') == 'agregarLivedata':
        debugPrint("Ingreso a POST agreegar livedata")
        ubicacion = request.POST.get('ubicacion')
        carid = request.POST.get('cardid')
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        cargo = request.POST.get('cargo')
        fecha = request.POST.get('f_ingreso')
        hora = request.POST.get('h_ingreso')
        if request.POST.get('cardid') and request.POST.get('nombre') and request.POST.get('apellido') and request.POST.get('cargo') and request.POST.get('f_ingreso') and request.POST.get('h_ingreso'):
            try:
                users = models.LiveData.objects.get(
                    card_id=request.POST.get('cardid'))
                datos = {'Estado': "Invalido",
                         'Mensaje': "El Card ID ingresado ya se encuentra al interior de la mina"}
                return JsonResponse(datos)
            except ObjectDoesNotExist:
                None
            cantidadactualRegistrada = models.LiveData.objects.all().count()
            user = models.LiveData()
            user.id = cantidadactualRegistrada+1
            user.card_id = carid
            user.nombre = nombre
            user.apellido = apellido
            user.cargo = cargo
            user.f_ingreso = fecha
            user.h_ingreso = hora
            user.save()
            debugPrint("GuardaLivedata")
            user2 = models.Historial()
            cantidadactualRegistrada = models.Historial.objects.all()
            debugPrint("Total: ", cantidadactualRegistrada)
            user2.id = cantidadactualRegistrada + 1
            user2.card_id = carid
            user2.nombre = nombre
            user2.apellido = apellido
            user2.cargo = cargo
            user2.f_evento = fecha
            user2.h_evento = hora
            user2.save()
            debugPrint("GuardaHistorial")
            try:
                users = models.PersonalRegistrado.objects.get(
                    nombre=request.POST.get('nombre'), apellido=request.POST.get('apellido'))
            except ObjectDoesNotExist:
                user3 = models.NoRegistrados()
                user3.ubicacion = ubicacion
                user3.card_id = carid
                user3.f_evento = fecha
                user3.h_evento = hora
                user3.evento = 'Ingreso'
                user3.save()
                debugPrint("Guarda no registrados")
            return JsonResponse({'Mensaje': "El personal ha sido agregado con exito"})
        datos = {"Estado": "Invalido",
                 'Mensaje': "Debe ingresar todos los campos correctamente"}
        return JsonResponse(datos)

    elif request.method == 'GET':
        return render(request, "livedata/livedata_agregar.html")


@login_required(login_url='autenticacion')
def livedata_eliminar(request):
    if request.method == "POST" and request.POST.get("Comando") == "ObtenerDatosCardID":
        debugPrint(request.POST.get('card_id'))
        if request.POST.get('card_id'):
            card_id_a_borrar = request.POST.get('card_id')
            tupla = models.LiveData.objects.get(card_id=card_id_a_borrar)
            return JsonResponse({"Estado": "Valido",
                                 "ubicacion": tupla.ubicacion,
                                 "cardid": tupla.card_id,
                                 "nombre": tupla.nombre,
                                 "apellido": tupla.ap_paterno,
                                 "cargo": tupla.cargo,
                                 "fechaingreso": tupla.f_ingreso,
                                 "horaingreso": tupla.h_ingreso
                                 })
        else:
            return JsonResponse({"Estado": "Invalido"})
    elif request.method == "POST" and request.POST.get("Comando") == "EliminarEntrada":
        if request.POST.get('card_id'):
            try:
                card_id_a_borrar = request.POST.get('card_id')
                tupla = models.LiveData.objects.get(
                    card_id=card_id_a_borrar)
                tupla.delete()
                return JsonResponse({"Estado": "Valido", "Mensaje": "El registro ha sido eliminado"})
            except Exception as ex:
                return JsonResponse({"Estado": "Invalido", "Mensaje": f"{str(ex)}"})
        return JsonResponse({"Estado": "Invalido", "Mensaje": "El CardID seleccionado no se encuentra en la base de datos"})
    if request.method == "GET":
        users = models.LiveData.objects.all()
        datos = {'livedata': users}
        return render(request, "livedata/livedata_eliminar.html", datos)


Status2String = {
    "0": "Agregado automatico",
    "1": "Agregado manual",
    "10": "Borrado de agregado automatico",
    "11": "Borrado de agregado manual",
}


@login_required(login_url='autenticacion')
def marcacion(request):
    if request.method == "POST" and request.POST.get("Comando") == "TablaMarcacion":
        debugPrint("Search")
        min = request.POST.get("min")
        max = request.POST.get("max")
        draw = int(request.POST.get('draw', 0))
        start = int(request.POST.get('start', 0))
        length = int(request.POST.get('length', 0))
        search_value = request.POST.get('search[value]', '')
        queryset = models.Historial.objects.order_by(
            '-f_evento', '-h_evento', '-id')
        if min != "" and max != "":
            queryset = queryset.filter(
                f_evento__range=(min, max)).order_by('-f_evento', '-h_evento', '-id')
        if search_value != '':
            queryset = queryset.filter(
                Q(id__icontains=search_value) |
                Q(ap_paterno__icontains=search_value) |
                Q(ap_materno__icontains=search_value) |
                Q(nombre__icontains=search_value) |
                Q(dni__icontains=search_value) |
                Q(proyecto__icontains=search_value) |
                Q(cargo__icontains=search_value) |
                Q(area__icontains=search_value) |
                Q(servicio__icontains=search_value) |
                Q(supervision__icontains=search_value) |
                Q(guardia__icontains=search_value) |
                Q(status__icontains=search_value)
            ).order_by('-f_evento', '-h_evento', '-id')
        total_records = queryset.count()
        queryset = queryset[start:start+length]
        data = []
        id = total_records - start + 1
        for obj in queryset:
            id -= 1
            item = {
                'primarykey': obj.id,
                'id': id,
                'card_id': obj.card_id,
                'nombre': obj.nombre,
                'ap_paterno': obj.ap_paterno,
                'ap_materno': obj.ap_materno,
                'dni': obj.dni,
                'proyecto': obj.proyecto,
                'cargo': obj.cargo,
                'area': obj.area,
                'servicio': obj.servicio,
                'supervision': obj.supervision,
                'guardia': obj.guardia,
                'f_evento': obj.f_evento,
                'h_evento': obj.h_evento,
                'evento': obj.evento,
                'status': Status2String[obj.status],
            }
            data.append(item)
        response = {
            "draw": draw,
            "recordsTotal": total_records,
            "recordsFiltered": total_records,
            "data": data
        }
        # debugPrint(response)
        return JsonResponse(response)
    elif request.method == "GET" and request.GET.get("Comando") == "DescargarExcel":
        min = request.GET.get('FechaInicial')
        max = request.GET.get('FechaFinal')
        search_value = request.GET.get('Search', '')
        debugPrint(f"{min}, {max}")
        queryset = models.Historial.objects.order_by('-id')
        if min != "" and max != "":
            queryset = queryset.filter(
                f_evento__range=(min, max)).order_by('-id')
        if search_value != '':
            queryset = queryset.filter(
                Q(id__icontains=search_value) |
                Q(ap_paterno__icontains=search_value) |
                Q(ap_materno__icontains=search_value) |
                Q(nombre__icontains=search_value) |
                Q(dni__icontains=search_value) |
                Q(proyecto__icontains=search_value) |
                Q(cargo__icontains=search_value) |
                Q(area__icontains=search_value) |
                Q(servicio__icontains=search_value) |
                Q(supervision__icontains=search_value) |
                Q(guardia__icontains=search_value) |
                Q(status__icontains=search_value)
            ).order_by("'-f_evento', '-h_evento', '-id'")
        total_records = queryset.count() + 1
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['#', 'N° PERSONA', 'APELLIDO PATERNO', 'APELLIDO MATERNO', 'NOMBRE',
                  'DNI', 'FECHA DE NACIMIENTO', 'PROYECTO', 'CENTRO DE COSTE', "TIPO DE TRABAJADOR",
                   "CLAVE DE SEXO", "FECHA DE ALTA", "FECHA DE BAJA", "MOTIVO DE CESE", "CARGO",
                   "CÓDIGO DE TARJETA", "ÁREA", "SERVICIO", "SUPERVISIÓN", "GUARDIA", "CORREO",
                   "N° CELULAR", "FECHA EVENTO", "HORA EVENTO", "EVENTO", "STATUS"])
        for item in queryset:
            total_records -= 1
            ws.append([total_records, item.n_persona, item.ap_paterno, item.ap_materno, item.nombre, item.dni, item.f_nac,
                       item.proyecto, item.centro_coste, item.tipo_trabajador, item.clave_sexo, item.f_alta,
                       item.f_baja, item.motivo_cese, item.cargo, item.card_id, item.area, item.servicio,
                       item.supervision, item.guardia, item.correo, item.n_celular, item.f_evento,
                       item.h_evento, item.evento, item.status])
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Historial_ingresos_salidas.xlsx"'
        wb.save(response)
        return response
    elif request.method == "POST" and request.POST.get("comando") == "eliminarRegistro":
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({"Estado": "noExitoso", "Mensaje": "Su cuenta no tiene los permisos para hacer esta accion"})
        debugPrint("ingreso a eliminar registro")
        primarykey = request.POST.get("primarykey")
        registroHistorial = models.Historial.objects.get(id=primarykey)
        if registroHistorial.status == '0':
            registroHistorial.status = "10"
        elif registroHistorial.status == '1':
            registroHistorial.status = "11"
        else:
            debugPrint("respuesta estado, mensaje no exitoso")
            return JsonResponse({"Estado": "noExitoso", "Mensaje": "El registro se encontraba eliminado previamente"})
        try:
            registroLivedata = models.LiveData.objects.get(
                card_id=registroHistorial.card_id)
            debugPrint("Se encontro en registro en LiveData")
            if registroHistorial.evento == "Ingreso" and registroHistorial.f_evento == registroLivedata.f_ingreso and registroHistorial.h_evento == registroLivedata.h_ingreso:
                ultimoHistorial = models.Historial.objects.filter(
                    Q(card_id=registroHistorial.card_id) & (Q(status='0') | Q(status='1'))).order_by(
                    '-f_evento', '-h_evento')
                debugPrint("El registro encontrado es ingreso")
                if ultimoHistorial[0] == registroHistorial:
                    try:
                        if ultimoHistorial[1].evento == "Salida":
                            registroLivedata.delete()
                    except:
                        registroLivedata.delete()
        except:
            debugPrint("Except al encontrar el livedata del registro eliminado")
        if registroHistorial.evento == "Salida":
            debugPrint("El registro eliminado es salida")
            ultimoHistorial = models.Historial.objects.filter(
                Q(card_id=registroHistorial.card_id) & (Q(status='0') | Q(status='1'))).order_by(
                '-f_evento', '-h_evento')
            if ultimoHistorial[0] == registroHistorial:
                debugPrint("Es el ultimo historial eliminado")
                debugPrint(len(ultimoHistorial))
                try:
                    if ultimoHistorial[1].evento == "Ingreso":
                        nuevoLivedata = models.LiveData()
                        nuevoLivedata.ubicacion = ultimoHistorial[1].ubicacion
                        nuevoLivedata.f_ingreso = ultimoHistorial[1].f_evento
                        nuevoLivedata.h_ingreso = ultimoHistorial[1].h_evento
                        nuevoLivedata.n_persona = ultimoHistorial[1].n_persona
                        nuevoLivedata.ap_paterno = ultimoHistorial[1].ap_paterno
                        nuevoLivedata.ap_materno = ultimoHistorial[1].ap_materno
                        nuevoLivedata.nombre = ultimoHistorial[1].nombre
                        nuevoLivedata.dni = ultimoHistorial[1].dni
                        nuevoLivedata.f_nac = ultimoHistorial[1].f_nac
                        nuevoLivedata.proyecto = ultimoHistorial[1].proyecto
                        nuevoLivedata.centro_coste = ultimoHistorial[1].centro_coste
                        nuevoLivedata.tipo_trabajador = ultimoHistorial[1].tipo_trabajador
                        nuevoLivedata.clave_sexo = ultimoHistorial[1].clave_sexo
                        nuevoLivedata.f_alta = ultimoHistorial[1].f_alta
                        nuevoLivedata.f_baja = ultimoHistorial[1].f_baja
                        nuevoLivedata.motivo_cese = ultimoHistorial[1].motivo_cese
                        nuevoLivedata.cargo = ultimoHistorial[1].cargo
                        nuevoLivedata.card_id = ultimoHistorial[1].card_id
                        nuevoLivedata.area = ultimoHistorial[1].area
                        nuevoLivedata.servicio = ultimoHistorial[1].servicio
                        nuevoLivedata.supervision = ultimoHistorial[1].supervision
                        nuevoLivedata.guardia = ultimoHistorial[1].guardia
                        nuevoLivedata.correo = ultimoHistorial[1].correo
                        nuevoLivedata.n_celular = ultimoHistorial[1].n_celular
                        nuevoLivedata.save()
                except:
                    debugPrint(
                        "Error al obtener el ultimo historial asociado al registro eliminado SALIDA")
                debugPrint("eliminado con registro agregado salida")
        debugPrint("responde eliminacion correcta")
        registroHistorial.save()
        return JsonResponse({"Mensaje": "El registro se ha eliminado correctamente"})
    debugPrint("Marcacion")
    # users = models.Historial.objects.all().order_by('-id')
    # datos = { 'marcacion' : users}
    # debugPrint(users[0].id, users[0].f_evento, type(users))
    return render(request, "marcacion/marcacion.html")


@login_required(login_url='autenticacion')
def marcacion_agregar(request):
    if not request.user.is_superuser and not request.user.is_staff:
        return render(request, "marcacion/marcacion_agregar_denegado.html")
    debugPrint(request.POST.get('comando'))
    if request.method == 'POST' and request.POST.get('comando') == 'consultaDatos':
        debugPrint(request.POST.get('cardid'))
        debugPrint(request.POST.get('nombre'))
        debugPrint(request.POST.get('apellido'))
        debugPrint(request.POST.get('cargo'))
        debugPrint(request.POST.get('f_ingreso'))
        debugPrint(request.POST.get('h_ingreso'))
        # agregar datos
        if len(request.POST.get('dni', ' ')) != 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El DNI debe ser de 8 caracteres"})
        elif not request.POST.get('nombre') and not request.POST.get('apellido') and not request.POST.get('cargo') and not request.POST.get('f_ingreso') and not request.POST.get('h_ingreso'):
            try:
                users = models.PersonalRegistrado.objects.get(
                    dni=request.POST.get('dni'))
            except ObjectDoesNotExist:
                return JsonResponse({"Estado": "Invalido", "Mensaje": "El DNI no se encuentra registrado en la planilla"})
            else:
                return JsonResponse({"card_id": users.card_id,
                                    "nombre": users.nombre,
                                     "apellido": users.ap_paterno,
                                     "cargo": users.cargo,
                                     "guardia": users.guardia,
                                     "f_ingreso": '',
                                     "h_ingreso": ''})

    if request.method == 'POST' and request.POST.get('comando') == 'agregarLivedata':
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({"Estado": "noExitoso", "Mensaje": "Su cuenta no tiene los permisos para hacer esta accion"})
        debugPrint("Ingreso a POST agreegar livedata")
        dni = request.POST.get('dni')
        try:
            users = models.PersonalRegistrado.objects.get(
                dni=request.POST.get('dni'))
        except ObjectDoesNotExist:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El DNI no se encuentra registrado en la planilla"})
        fecha = request.POST.get('f_ingreso')
        hora = request.POST.get('h_ingreso')
        evento = request.POST.get("evento")
        if fecha and hora and evento:
            user2 = models.Historial()
            cantidadactualRegistrada = models.Historial.objects.all().count()
            user2.id = cantidadactualRegistrada + 1
            user2.ubicacion = "San Rafael"
            user2.f_evento = fecha
            user2.h_evento = hora
            user2.evento = evento
            user2.status = "1"
            user2.n_persona = users.n_persona
            user2.ap_paterno = users.ap_paterno
            user2.ap_materno = users.ap_materno
            user2.nombre = users.nombre
            user2.dni = users.dni
            user2.f_nac = users.f_nac
            user2.proyecto = users.proyecto
            user2.centro_coste = users.centro_coste
            user2.tipo_trabajador = users.tipo_trabajador
            user2.clave_sexo = users.clave_sexo
            user2.f_alta = users.f_alta
            user2.f_baja = users.f_baja
            user2.motivo_cese = users.motivo_cese
            user2.cargo = users.cargo
            user2.card_id = users.card_id
            user2.area = users.area
            user2.servicio = users.servicio
            user2.supervision = users.supervision
            user2.guardia = users.guardia
            user2.correo = users.correo
            user2.n_celular = users.n_celular
            user2.save()
            debugPrint("GuardaHistorial")
            ultimoHistorial = models.Historial.objects.filter(
                Q(dni=dni) & (Q(status='0') | Q(status='1'))).order_by('-f_evento', '-h_evento', 'id')
            if ultimoHistorial[0] != user2:
                debugPrint(
                    "Registro guardado pero no es el ultimo del historial")
                return JsonResponse({"Mensaje": "El registro se ha agregado correctamente"})
            if user2.evento == "Salida":
                try:
                    userLiveData = models.LiveData.objects.get(dni=dni)
                    userLiveData.delete()
                    debugPrint("Se elimino registro de livedata")
                    return JsonResponse({"Mensaje": "Registro agregado correctamente..."})
                except:
                    debugPrint(
                        "Registro guardado pero debio eliminarse de livedata pero no se pudo")
                    return JsonResponse({"Mensaje": "Registro agregado correctamente..."})
            try:
                userLiveData = models.LiveData.objects.get(dni=dni)
                userLiveData.f_ingusersuserreso = fecha
                userLiveData.h_ingreso = hora
                userLiveData.save()
            except:
                userLiveData = models.LiveData()
                user2.ubicacion = "San Rafael"
                userLiveData.f_ingreso = fecha
                userLiveData.h_ingreso = hora
                userLiveData.n_persona = users.n_persona
                userLiveData.ap_paterno = users.ap_paterno
                userLiveData.ap_materno = users.ap_materno
                userLiveData.nombre = users.nombre
                userLiveData.dni = users.dni
                userLiveData.f_nac = users.f_nac
                userLiveData.proyecto = users.proyecto
                userLiveData.centro_coste = users.centro_coste
                userLiveData.tipo_trabajador = users.tipo_trabajador
                userLiveData.clave_sexo = users.clave_sexo
                userLiveData.f_alta = users.f_alta
                userLiveData.f_baja = users.f_baja
                userLiveData.motivo_cese = users.motivo_cese
                userLiveData.cargo = users.cargo
                userLiveData.card_id = users.card_id
                userLiveData.area = users.area
                userLiveData.servicio = users.servicio
                userLiveData.supervision = users.supervision
                userLiveData.guardia = users.guardia
                userLiveData.correo = users.correo
                userLiveData.n_celular = users.n_celular
                userLiveData.save()
            debugPrint("El registro se agrego y se actualizo livedata")
            return JsonResponse({'Mensaje': "El registro se ha agregado correctamente."})
        datos = {"Estado": "Invalido",
                 'Mensaje': "Debe ingresar todos los campos correctamente"}
        return JsonResponse(datos)

    elif request.method == 'GET':
        return render(request, "marcacion/marcacion_agregar.html")


@login_required(login_url='autenticacion')
def noregistrados(request):
    users = models.NoRegistrados.objects.all()
    datos = {'noregistrados': users}
    return render(request, "noregistrados/noregistrados.html", datos)


@login_required(login_url='autenticacion')
def registrarusuario(request):
    if request.method == "POST":
        if request.POST.get("Comando") != "RegistrarUsuario":
            return JsonResponse({"Estado": "Invalido", "Mensaje": "¿Que haces?"})
        Nombre = request.POST.get("Nombre").strip().upper()
        PrimerApellido = request.POST.get("PrimerApellido").strip().upper()
        SegundoApellido = request.POST.get("SegundoApellido").strip().upper()
        DNI = request.POST.get("DNI").strip()
        Correo = request.POST.get("Correo").strip()
        Telefono = request.POST.get("Telefono").strip()
        is_staff = False
        if request.POST.get("Rol") == "Observador y registro de planilla":
            is_staff = True
        if not (len(Nombre) and len(PrimerApellido) and len(SegundoApellido) and len(DNI) and len(Correo) and len(Telefono)):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "Debe llenar todos los campos antes de continuar"})
        if not re.match(r'^[A-Za-z\s]+$', Nombre):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El nombre debe contener solo letras"})
        if not re.match(r'^[A-Za-z]+$', PrimerApellido):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El apellido debe contener solo letras"})
        if not re.match(r'^[A-Za-z\s]+$', SegundoApellido):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El apellido debe contener solo letras"})
        if len(DNI) != 8:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El DNI debe tener 8 caracteres"})
        if not re.match(r'^\d+$', DNI):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "Debe ingresar un dni valido"})
        if models.UserInfo.objects.filter(DNI=DNI).exists():
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El DNI ya se encuentra registrado"})
        if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', Correo):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "Debe ingresar un correo valido"})
        if User.objects.filter(email=Correo).exists():
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El correo ya se encuentra registrado"})
        if len(Telefono) != 9:
            return JsonResponse({"Estado": "Invalido", "Mensaje": "El telefono debe tener 9 caracteres"})
        if not re.match(r'^\d+$', Telefono):
            return JsonResponse({"Estado": "Invalido", "Mensaje": "Debe ingresar un telefono valido"})
        try:
            Username = Nombre[0].lower() + PrimerApellido.lower() + DNI[4:8]
            if User.objects.filter(username=Username).exists():
                Username = Nombre[0].lower() + \
                    PrimerApellido.lower() + DNI[0:4]
            random.seed(int(datetime.now().timestamp()))
            Password = ''.join(random.choice(
                string.ascii_letters + string.digits) for _ in range(10))
            NuevoUsuario = User(username=Username, email=Correo,
                                first_name=Nombre, last_name=PrimerApellido, is_staff=is_staff)
            NuevoUsuario.set_password(Password)
            MensajeHTML = f"""\
            <html>
            <head></head>
            <body>
                <p>Hola, <span style="font-size: larger;">{NuevoUsuario.first_name}</span>!</p>
                <p>Se ha registrado su cuenta, sus datos de acceso son:</p>
                <p><br><span style="font-size: larger;"><b>Usuario: {NuevoUsuario.username}</b></span><br>
                <p><br><span style="font-size: larger;"><b>Contraseña: {Password}</b></span></p>
            </body>
            </html>
            """
            Fun.EnviaCorreo(
                NuevoUsuario.email, "Cuenta registrada - Control de acceso", MensajeHTML)
            NuevoUsuario.save()
            NuevoUsuarioInfo = models.UserInfo(
                User=NuevoUsuario, DNI=DNI, Telefono=Telefono, SegundoApellido=SegundoApellido)
            NuevoUsuarioInfo.save()
            debugPrint(f'{Username}, {Password}')
            return JsonResponse({"Estado": "Valido", "Mensaje": "Registrado correctamente, se ha enviado las credenciales al correo electronico"})
        except Exception as e:
            return JsonResponse({"Estado": "Invalido", "Mensaje": f"{str(e)}"})

    else:
        if request.user.is_superuser:
            return render(request, "registrarusuario/plantillaregistro.html")
        return render(request, "registrarusuario/plantilladenegado.html")


def eliminarusuario(request):
    if request.method == "POST":
        if request.POST.get("Comando") == "ConsultarDatos":
            Usuario = request.POST.get("Usuario")
            if not User.objects.filter(username=Usuario).exists():
                return JsonResponse({"Estado": "Invalido"})
            user = User.objects.get(username=Usuario)
            user2 = models.UserInfo.objects.get(User=user)
            Nombre = user.first_name
            PrimerApellido = user.last_name
            SegundoApellido = user2.SegundoApellido
            DNI = user2.DNI
            Correo = user.email
            Telefono = user2.Telefono
            Rol = "Solo observador"
            if user.is_staff:
                Rol = "Observador y registro de planilla"
            Data = {
                "Estado": "Valido",
                "Nombre": Nombre,
                "PrimerApellido": PrimerApellido,
                "SegundoApellido": SegundoApellido,
                "DNI": DNI,
                "Correo": Correo,
                "Telefono": Telefono,
                "Rol": Rol,
            }
            return JsonResponse(Data)
        elif request.POST.get("Comando") == "EliminarUsuario":
            try:
                Usuario = request.POST.get("Usuario")
                user = User.objects.get(username=Usuario)
                user.delete()
                return JsonResponse({"Estado": "Valido", "Mensaje": "El usuario se elimino correctamente"})
            except Exception as e:
                return JsonResponse({"Estado": "Invalido", "Mensaje": f"{str(e)}"})

    if request.user.is_superuser:
        Usuarios = User.objects.exclude(username=request.user.username)
        return render(request, "eliminarusuario/plantillaeliminar.html", {"Usuarios": Usuarios})
    return render(request, "eliminarusuario/plantilladenegado.html")
